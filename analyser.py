from openpyxl import *
import glob
import sys

if len(sys.argv) < 3:
  print "Not enough arguments: inputdir and outputdir required"
inputdir = sys.argv[1]
outputdir = sys.argv[2]

inputfiles = glob.glob(inputdir + "/*.xlsx")

def verify_prereqs():
  ws = wb['Beat-To-Beat']
  if ws["E2"].value == "":
    print "Height missing in cell E2"
    exit
  if ws["F2"] == "":
    print "Weight missing in cell F2"
    exit

def step_4():
  ws = wb['Beat-To-Beat']
  for i in [13,12,11,10,9,2]:
    delete_column(ws, i)

def step_6():
  copy_cells('Cardiac Parameters', [9,10,11,13,14], 'Beat-To-Beat', 7)

def step_8():
  copy_cells('HRV', [3,4,6,7,8], 'Beat-To-Beat', 12)

def step_10():
  copy_cells('BPV (dia.)', [3,4,6,7,8], 'Beat-To-Beat', 17)

def step_11():
  ws = wb['Beat-To-Beat']
  col = ws.max_column + 1
  for row in range(1,ws.max_row + 1):
    ws.cell(row=row, column=col).value = ws.cell(row=row, column=2).value
  delete_column(ws, 2)

def step_12():
  ws = wb['Beat-To-Beat']
  ws["A3"].value = "BSA_m2"
  ws["B3"] = "=0.0235*E2^0.42246*F2^0.51456"

def step_13():
  ws = wb['Beat-To-Beat']
  for i in range(5):
    insert_column(ws, 7)
  headers = ["SI", "CO", "CI", "TPR", "TPRI"]
  
  for column in range(7,12):
    ws.cell(row=4, column=column).value = headers[column-7]
    for row in range(9, ws.max_row+1):
      val = ""
      if column == 7:
        val = "=F" + str(row) + "/$B$3"
      elif column == 8:
        val = "=B" + str(row) + "*F" + str(row) + "/1000"
      elif column == 9:
        val = "=H" + str(row) + "/$B$3"
      elif column == 10:
        val = "=E" + str(row) + "/H" + str(row)
      elif column == 11:
        val = "=J" + str(row) + "/$B$3" 
      ws.cell(row=row, column=column).value = val

def step_14():
  ws = wb['Beat-To-Beat']
  insert_column(ws, 21)
  ws["U4"] = "LF_HF"
  for row in range(9, ws.max_row + 1):
    ws.cell(row=row, column=21).value = "=S" + str(row) + "/T" + str(row)

def step_14_2():
  ws = wb['Beat-To-Beat']
  ws["AC4"] = "Diff_RRI"
  ws["AD4"] = "Diff_RRIsqr"
  ws["AE4"] = "Diff50_yes"
  for row in range(9, ws.max_row + 1):
    ws.cell(row=row, column=29).value = "=AB" + str(row+1) + "-AB" + str(row)
    ws.cell(row=row, column=30).value = "=AC" + str(row) + "^2"
    ws.cell(row=row, column=31).value = "=IF(AD" + str(row) + ">2500,1,0)"

def step_17():
  ws = wb['Beat-To-Beat']
  for column in range(2,28):
    ws.cell(row=4, column=column+30).value = ws.cell(row=4, column=column).value
  ws["BG4"] = "SDNN"
  ws["BH4"] = "MSD"
  ws["BI4"] = "R_MSD"
  ws["BJ4"] = "Sum_50"
  ws["BK4"] = "Count_50"
  ws["BL4"] = "pNN50"

def step_18():
  ws = wb['Beat-To-Beat']
  rows = find_window_rows(find_windows())
  make_medians(rows, ws, 'rec_s', 9)
  make_hrv(rows, ws, 'rec_s', 9)
  make_medians(rows, ws, 'met_s', 10)
  make_hrv(rows, ws, 'met_s', 10)
  make_medians(rows, ws, 'sta_s', 11)
  make_hrv(rows, ws, 'sta_s', 11)

def make_hrv(rows, ws, name, row):
  start_row = str(rows[name][0])
  stop_row = str(rows[name][1])
  ws["BG" + str(row)] = "=STDEV.P(AB" + start_row + ":AB" + stop_row + ")"
  ws["BH" + str(row)] = "=AVERAGE(AD" + start_row + ":AD" + stop_row + ")"
  ws["BI" + str(row)] = "=SQRT(BH" + str(row) + ")"
  ws["BJ" + str(row)] = "=SUM(AE" + start_row + ":AE" + stop_row + ")"
  ws["BK" + str(row)] = "=COUNT(AE" + start_row + ":AE" + stop_row + ")"
  ws["BL" + str(row)] = "=BJ" + str(row) + "/BK" + str(row)

def make_medians(rows, ws, name, row):
  for column in range(2,29):
    char = utils.cell.get_column_letter(column)
    start_row = str(rows[name][0])
    stop_row = str(rows[name][1])
    ws.cell(row=row, column=column+30).value = "=MEDIAN(" + char + start_row + ":" + char + stop_row + ")"

def find_window_rows(windows):
  ws = wb['Beat-To-Beat']
  results = dict()
  started = windows.keys()
  for row in range(9, ws.max_row + 1):
    time = ws.cell(row=row, column=1).value
    for key, val in windows.iteritems():
      if time > val[0] and time < val[1] and key in started:
        results[key] = row
        started.remove(key)
        print key
      elif time > val[1] and type(results[key]) is int:
        results[key] = (results[key], row)
        print key
  return results
  

def find_windows():
  ws = wb['HRV']
  results = dict()
  for row in range(6, ws.max_row + 1):
    name = ws.cell(row=row, column=2).value
    if name and type(name) is unicode:
      time = ws.cell(row=row, column=1).value
      if "ceba metronom" in name:
        results["met_s"] = (time + 30, time + 240 + 30)
      elif "Start Recording" in name:
        results["rec_s"] = (time + 30, time + 240 + 30)
      elif "staaende" in name:
        print "STAAENDE " + str(row)
        results["sta_s"] = (time + 30, time + 120 + 30)
  return results

def copy_cells(ws1, ws1_columns, ws2, ws2_column_start):
  ws_c = wb[ws1]
  ws_b = wb[ws2]
  i = ws2_column_start
  for column in ws1_columns:
    i += 1
    for row in range(3, ws_c.max_row+1):
      ws_b.cell(row=row, column=i).value = ws_c.cell(row=row, column=column).value

def insert_column(ws, insert_column):
  for column in reversed(range(insert_column, ws.max_column + 2)):
    for row in range(3, ws.max_row + 1):
      try:
        prev_val = ws.cell(row=row, column=column-1).value
        ws.cell(row=row, column=column).value = prev_val
      except:
        print "Error at row " + str(row) + " and column " + str(column)
        raise

  for row in range(1, ws.max_row + 1):
    ws.cell(row=row, column=insert_column).value = ""

def delete_column(ws, delete_column):
  for column in range(delete_column, ws.max_column + 1):
    for row in range(1, ws.max_row + 1):
      ws.cell(row=row, column=column).value = ws.cell(row=row, column=column+1).value

for in_file in inputfiles:
  print "Working on file " + in_file
  out_file = "out." + os.path.basename(in_file)
  wb = load_workbook(in_file)
  verify_prereqs()
  step_4()
  step_6()
  step_8()
  step_10()
  step_11()
  step_12()
  step_13()
  step_14()
  step_14_2()
  step_17()
  step_18() 
  wb.save(outputdir + "/" + out_file)
